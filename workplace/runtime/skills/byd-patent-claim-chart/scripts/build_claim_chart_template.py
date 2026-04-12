from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN = Side(style="thin", color="D9DEE8")
HEADER_FILL = PatternFill("solid", fgColor="D69A3C")
SUB_FILL = PatternFill("solid", fgColor="F7E8C7")
META_FILL = PatternFill("solid", fgColor="F5F7FA")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _apply_border(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = border


def build_template(output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "专利侵权分析"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A14"

    widths = {
        "A": 14,
        "B": 12,
        "C": 30,
        "D": 22,
        "E": 28,
        "F": 24,
        "G": 28,
        "H": 12,
        "I": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[13].height = 34

    ws.merge_cells("A1:I1")
    ws["A1"] = "比亚迪中文专利竞品技术特征对比表"
    ws["A1"].font = Font(size=16, bold=True, color="1F2430")
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:I2")
    ws["A2"] = "用途：针对比亚迪有权中文专利，基于公开证据对竞品进行逐条权利要求技术特征比对"
    ws["A2"].font = Font(size=10, color="556070")
    ws["A2"].alignment = CENTER

    ws.merge_cells("A3:I3")
    ws["A3"] = "边界：本表用于公开证据比对和风险初筛，不直接作出法律侵权结论"
    ws["A3"].font = Font(size=10, color="8B90A0", italic=True)
    ws["A3"].alignment = CENTER

    meta_rows = [
        ("A5", "专利号", "B5", ""),
        ("D5", "专利名称", "E5", ""),
        ("H5", "报告日期", "I5", ""),
        ("A6", "专利权人", "B6", "比亚迪"),
        ("D6", "分析权利要求", "E6", ""),
        ("H6", "竞品数量", "I6", ""),
        ("A7", "竞品名称", "B7", ""),
        ("D7", "竞品型号", "E7", ""),
        ("H7", "分析人", "I7", ""),
        ("A8", "分析方法", "B8", "公开证据逐条比对"),
        ("D8", "判断口径", "E8", "明确匹配/部分匹配/可能匹配/证据不足/明显不匹配"),
        ("H8", "备注", "I8", ""),
    ]
    for label_cell, label, value_cell, value in meta_rows:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, color="344054")
        ws[label_cell].fill = META_FILL
        ws[label_cell].alignment = CENTER
        ws[value_cell] = value
        ws[value_cell].alignment = LEFT

    ws.merge_cells("B5:C5")
    ws.merge_cells("E5:G5")
    ws.merge_cells("B6:C6")
    ws.merge_cells("E6:G6")
    ws.merge_cells("B7:C7")
    ws.merge_cells("E7:G7")
    ws.merge_cells("B8:C8")
    ws.merge_cells("E8:G8")

    summary_labels = [
        ("B10", "明确匹配"),
        ("D10", "部分匹配"),
        ("F10", "可能匹配"),
        ("H10", "证据不足"),
        ("B11", "明显不匹配"),
    ]
    for cell, text in summary_labels:
        ws[cell] = text
        ws[cell].fill = SUB_FILL
        ws[cell].font = Font(bold=True, color="7A4D00")
        ws[cell].alignment = CENTER
    for cell in ["C10", "E10", "G10", "I10", "C11"]:
        ws[cell] = 0
        ws[cell].alignment = CENTER

    headers = [
        "权利要求",
        "特征编号",
        "权利要求技术特征",
        "竞品/型号",
        "竞品对应特征",
        "证据摘要",
        "证据链接",
        "匹配判断",
        "备注",
    ]
    for idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=13, column=idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = CENTER

    sample_rows = [
        [
            "权利要求1",
            "1.1",
            "示例：拆分后的单一技术特征",
            "示例竞品A / 型号A1",
            "基于公开资料提取的竞品对应特征",
            "证据摘要：说明该公开资料如何支持本项特征",
            "https://example.com/evidence-1",
            "明确匹配",
            "如有推断或争议点，在此说明",
        ],
        [
            "权利要求1",
            "1.2",
            "示例：另一条技术特征",
            "示例竞品A / 型号A1",
            "如无明确信息，可写证据不足",
            "证据摘要：资料未明示，仅能部分支持",
            "https://example.com/evidence-2",
            "证据不足",
            "不要把推断写成明示证据",
        ],
    ]
    start_row = 14
    for offset, row in enumerate(sample_rows):
        for idx, value in enumerate(row, start=1):
            cell = ws.cell(row=start_row + offset, column=idx, value=value)
            cell.alignment = LEFT if idx not in {1, 2, 8} else CENTER

    _apply_border(ws, 5, 8, 1, 9)
    _apply_border(ws, 10, 11, 2, 9)
    _apply_border(ws, 13, 15, 1, 9)

    for row in range(14, 80):
        ws.row_dimensions[row].height = 36

    ws["A10"] = "结论统计"
    ws["A10"].font = Font(bold=True, color="344054")
    ws["A10"].alignment = CENTER
    ws["A11"] = "补充统计"
    ws["A11"].font = Font(bold=True, color="344054")
    ws["A11"].alignment = CENTER

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    skill_root = Path(__file__).resolve().parents[1]
    output = skill_root / "assets" / "byd_patent_claim_chart_template.xlsx"
    build_template(output)
    print(output)


if __name__ == "__main__":
    main()
